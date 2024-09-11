# encoding: utf-8
############################################
#本文件包括两个功能：
#1、通过excel文件生成csv文件
#2、根据excel文件头描述生成cs文件
############################################
import openpyxl
import sys
import os
import json
import ConvertCommonType
from py.userType import USERTYPE

#代码模板文件名
TEMPLATE_FILE_NAME="templates/ResTemplate.cs"
TEMPLATE_LINE_NAME="templates/ResLineBaseTemplate.cs"
TEMPLATE_ENUM_NAME = "templates/EnumTemplate.cs"
TEMPLATE_TEXT_NAME = "templates/TextTemplate.cs"
TEMPLATE_LETTER_NAME="templates/ResLetterBaseTemplate.cs"

BASETYPE = ["string","int","float","short","bool"]

covData = False
covCode = False

def readExcel(path):
    wb=openpyxl.load_workbook(path, data_only=True)
    for wsname in wb.sheetnames:
        ws = wb[wsname]
        type = getDataSheetType(ws);
        dec = str(ws.cell(row=1, column=6).value)
        name = str(ws.cell(row=1, column=4).value)
        
        maxRow = ws.max_row+1
        maxCol = ws.max_column+1

        #添加新的模板生成  -------------   data 和line
        if type == "data" or type == "line" or type == "letter":
            
            colList = []   #需要转表列的标记
            fieldDict = {} #字段-类型字典
            fieldList = [] #字段数组
            arrayDict = {} #数组字段-长度字典
            userTypeDict = {} #用户类型-长度字典
            shortTypeArray = [] #简易数组字段 (用[]方式定义的数组)
            
            lastTypeName = ""
            lastFieldName = ""
            
            #先剔除掉辅助列
            for colIndex in range(1,maxCol):            
                typename = ws.cell(row=2, column=colIndex).value
                fieldname = ws.cell(row=4, column=colIndex).value
                
                #当字段名不为空时，认为是数据，否则都认为是注释
                if fieldname != None and typename != "des":
                    flagIndex = fieldname.find('#')
                    if flagIndex >= 0 :
                        fieldname = fieldname[0:flagIndex]
                    isUserType = False
                    
                    #如果字段名有'.'，说明是用户类型
                    if fieldname.find('.') >= 0:
                        isUserType = True
                        fieldname = fieldname.split('.')[0]
                        if fieldname in userTypeDict and typename == None:
                            userTypeDict[fieldname] += 1;
                        elif typename != None:
                            userTypeDict[fieldname] = 1;

                    if not isUserType or (isUserType == True and typename != None):
                        #当类型和字段名与前一个字段相同，则认为是数组
                        if lastTypeName == typename and lastFieldName == fieldname:
                            if fieldname in arrayDict:
                                arrayDict[fieldname] += 1;
                            else:
                                arrayDict[fieldname] = 2;

                        #如果名字段名有[],则认为是数组
                        elif typename.find("[") >=0:
                            defArr = typename.split('[')[0]
                            #typename = defArr[0]
                            if defArr in BASETYPE:
                                fieldDict[fieldname] = typename
                                fieldList.append(fieldname)
                                #arrayDict[fieldname] = int(defArr[1][0:-1])
                                shortTypeArray.append(fieldname)
                        #不满足以上条件，则认为是普通数据
                        else:
                            fieldDict[fieldname] = typename
                            fieldList.append(fieldname)


                    if isUserType != True:
                        lastTypeName = str(typename)
                        lastFieldName = str(fieldname)
                    elif isUserType == True and typename != None:
                        lastTypeName = typename
                        lastFieldName = fieldname
                    colList.append(colIndex)

        #print(fieldDict)
        #print(fieldList)
        #print(arrayDict)
        #print(userTypeDict)
            
            #读取excel文件并生成CSV文件
            if covData:
                file = open('out/data/'+name+'.csv', 'w', encoding='utf-8')
                for rowIndex in range(5,maxRow):
                    id = ws.cell(row=rowIndex, column=1).value
                    if id == None:
                        continue
                    line = "";
                    for colIndex in colList:
                        data = ws.cell(row=rowIndex, column=colIndex).value
                        if data != None:
                            line += str(ws.cell(row=rowIndex, column=colIndex).value).strip()
                        else:
                            #dataType = fieldDict[fieldList[colIndex-1]]
                            #line += DEFAULT_TYPE_VALUE_DICT[dataType]
                            line += ''
                        line += '|'
                    writeCSVFile(file,line,rowIndex != maxRow - 1)
                file.close()

            #通过模板生成C#文件
            if covCode:
                attr = ""
                method = ""
                index = 0
                hasKey = "false"
                
                for field in fieldList:
                    #简易数组
                    if field in shortTypeArray:
                        #print(field)
                        if fieldDict[field] != "string[]":
                            attr += "    public {0} {1};\n".format(fieldDict[field],field)
                        else:
                            attr += "    public {0} {1};\n".format(fieldDict[field],field)
                            attr += "    public {0} {1};\n".format("Dictionary<Language, string[]>",field+"LanguageArray")

                        if fieldDict[field] != "string[]":
                            method +="        ResourceInitUtil.InitShortArray(array[{0}],out this.{1});\n".format(index,field)
                        else:
                            method +="        ResourceInitUtil.InitShortArray(array[{0}],out this.{1},out this.{2});\n".format(index,field, field+"LanguageArray")
                        index+=1
                    #处理数组
                    elif field in arrayDict:
                        attr += "    public {0}[] {1};\n".format(fieldDict[field],field)
                        
                        if fieldDict[field] in USERTYPE:
                            UfieldName = fieldDict[field]
                            method += "        ResourceInitUtil.InitUserArray<{0}>(array,{1},{2},{3},out this.{4});\n".format(UfieldName,index,index+arrayDict[field]*userTypeDict[field]-1,userTypeDict[field],field)
                        else:
                            method += "        ResourceInitUtil.InitDefaultArray(array,{0},{1}, out this.{2});\n".format(index,index+arrayDict[field]-1,field)

                        arrayLength = 1
                        userTypeLength = 1
                        if field in arrayDict:
                            arrayLength = arrayDict[field]
                        if field in userTypeDict:
                            userTypeLength = userTypeDict[field]
                        index+=arrayLength * userTypeLength
                    #处理用户类型
                    elif fieldDict[field] in USERTYPE:
                        UfieldName = fieldDict[field]
                        attr += "    public {0} {1};\n".format(fieldDict[field],field)
                        method +="        ResourceInitUtil.InitUserField<{0}>(array,{1},{2},{3},out this.{4});\n".format(UfieldName,index,index+userTypeDict[field]-1,userTypeDict[field],field)
                        index += userTypeDict[field]
                    #处理枚举
                    elif fieldDict[field].startswith("e_"):
                        EfieldName = fieldDict[field].split("_")[1]
                        attr += "    public {0} {1};\n".format(EfieldName,field)
                        method += "        {0} = ({1})int.Parse(array[{2}]);\n".format(field,EfieldName,index)
                        index+=1
                    #处理非数组
                    else:
                        if field == "ID":
                            hasKey = "true"
                        elif fieldDict[field] == "string":
                            attr += "    public {0} {1};\n".format(fieldDict[field],field)
                            attr +=  "    public {0} {1};\n".format("Dictionary<Language, string>",field+"Language")
                        else:
                            attr += "    public {0} {1};\n".format(fieldDict[field],field)
                        if fieldDict[field] != "string":
                            method +="        ResourceInitUtil.InitField(array[{0}],out this.{1});\n".format(index,field)
                        else:
                             method +="        ResourceInitUtil.InitField(array[{0}],out this.{1},out this.{2});\n".format(index,field,field+"Language")
                        index += 1
                #------------- -----另外在这里特殊判断 新增的模板 ------------------------
                if type == "data":
                    writeCSharpFile(dec,name,attr,method,hasKey)
                elif type == "line":
                    writeCSharpLineFile(dec,name)
                elif type == "letter":
                    writeCSharpLetterFile(dec,name)
                
        #枚举
        if type == "enum":
            #只有生成代码时才转换枚举
            if covCode:
                context = ""
                for rowIndex in range(3,maxRow):
                    context += "        "
                    context += str(ws.cell(row=rowIndex, column=2).value) + " = " + str(ws.cell(row=rowIndex, column=1).value)
                    context += ", //"
                    context += str(ws.cell(row=rowIndex, column=3).value)
                    context += "\n"
                enumstr = getEnumString(name,context,dec)
                writeCSharpEnumFile(name,enumstr)         

        #分表
        if type == "addition":
            colList = []   #需要转表列的标记
            for colIndex in range(1,maxCol):
                typename = ws.cell(row=2, column=colIndex).value
                fieldname = ws.cell(row=4, column=colIndex).value
                
                #当字段名不为空时，认为是数据，否则都认为是注释
                if fieldname != None and typename != "des":
                    colList.append(colIndex)

            if covData:
                file = open('out/data/'+name+'.csv', 'a', encoding='utf-8')
                file.writelines('\n')
                for rowIndex in range(5,maxRow):
                    id = ws.cell(row=rowIndex, column=1).value
                    if id == None:
                        continue
                    line = "";
                    for colIndex in colList:
                        data = ws.cell(row=rowIndex, column=colIndex).value
                        if data != None:
                            line += str(ws.cell(row=rowIndex, column=colIndex).value).strip()
                        else:
                            #dataType = fieldDict[fieldList[colIndex-1]]
                            #line += DEFAULT_TYPE_VALUE_DICT[dataType]
                            line += ''
                        line += '|'
                    writeCSVFile(file,line,rowIndex != maxRow - 1)
                file.close()
        #单值
        if type == "keyValue":
            if covCode:
                context = ""
                for rowIndex in range(5,maxRow):
                    valuefieldType = str(ws.cell(row=rowIndex, column=4).value)
                    
                    if(valuefieldType.find('[]') < 0):
                        context += "        public const " + str(ws.cell(row=rowIndex, column=4).value) + " "
                    else:
                        context += "        public readonly static " + str(ws.cell(row=rowIndex, column=4).value) + " "
                    
                    context += str(ws.cell(row=rowIndex, column=1).value) + " = "
                    
                    if valuefieldType == "string":
                        context += "\"" + str(ws.cell(row=rowIndex, column=3).value) + "\""
                    elif valuefieldType == "float":
                        context += str(ws.cell(row=rowIndex, column=3).value) + "f"
                    elif valuefieldType == "bool":
                        if str(ws.cell(row=rowIndex, column=3).value).strip() == "0" :
                            context += "true"
                        else:
                            context += "false"
                    else:
                        context += str(ws.cell(row=rowIndex, column=3).value)
                    
                    context += "; //"
                    context += str(ws.cell(row=rowIndex, column=2).value)
                    context += "\n"
                enumstr = getValueString(name,context,dec)
                writeCSharpValueFile(name,enumstr)
        #文本 
        if type == "text":
            colList = []   #需要转表列的标记
            for colIndex in range(1,maxCol):
                typename = ws.cell(row=2, column=colIndex).value
                fieldname = ws.cell(row=4, column=colIndex).value
                
                #当字段名不为空时，认为是数据，否则都认为是注释,当字段名为name是，是描述字段，也不处理
                if fieldname != None and typename != "des" and typename != "textName":
                    colList.append(colIndex)

            context = "";

            if covData:
                isnewFile = True
                if os.path.exists('out/data/'+name+'.csv'):
                    isnewFile = False
                
                file = open('out/data/'+name+'.csv', 'a', encoding='utf-8')
                if not isnewFile :
                    file.writelines('\n')
                for rowIndex in range(5,maxRow):
                    id = ws.cell(row=rowIndex, column=1).value
                    if id == None:
                        continue
                    line = "";

                    context += "    public const int " + str(ws.cell(row=rowIndex, column=2).value) + " = " + str(ws.cell(row=rowIndex, column=1).value) + "; //" + str(ws.cell(row=rowIndex, column=3).value) + "\n"
                    for colIndex in colList:
                        data = ws.cell(row=rowIndex, column=colIndex).value
                        if data != None:
                            line += str(ws.cell(row=rowIndex, column=colIndex).value).strip()
                        else:
                            line += ''
                        line += '|'
                    writeCSVFile(file,line,rowIndex != maxRow - 1)
                file.close()

            file = open('out/tmp/'+name+'.txt', 'a', encoding='utf-8')
            file.write(context)
            file.close()


            file = open('out/tmp/'+name+'.txt', 'r', encoding='utf-8')
            dataString = ""
            try:
                dataString = file.read()
            finally:
                file.close()

            csTemplate = open(TEMPLATE_TEXT_NAME, 'r', encoding='utf-8')
            fileString = ""
            try:
                fileString = csTemplate.read()
            finally:
                csTemplate.close()

            csTarget = open('out/code/ResValue/'+name+".cs", 'w', encoding='utf-8')
            fileString = fileString.replace('{0}', dataString)

            csTarget.write(fileString)
            csTarget.close()

#获取枚举字符串
def getEnumString(name,context,des):
    enumString = "    /// <summary>\n"+"    /// "+ des +"\n"+"    /// </summary>\n"+"    public enum "+ name +"\n"+"    {"+"\n"+context+"    }\n"
    return enumString

#通过模板生成C# 枚举文件
def writeCSharpEnumFile(name,enumString):
    csTemplate = open(TEMPLATE_ENUM_NAME, 'r')
    fileString = ""
    try:
        fileString = csTemplate.read()
    finally:
        csTemplate.close()
        
    fileString = fileString.replace('{0}', enumString)
    
    csTarget = open('out/code/ResEnum/'+name+".cs", 'w')
    csTarget.write(fileString)
    csTarget.close()

def getValueString(name,context,des):
    enumString = "    /// <summary>\n"+"    /// "+ des +"\n"+"    /// </summary>\n"+"    public static class "+ name +"\n"+"    {"+"\n"+context+"    }\n"
    return enumString

#通过模版生产C# 单值文件
def writeCSharpValueFile(name, enumString):
    csTemplate = open(TEMPLATE_ENUM_NAME, 'r')
    fileString = ""
    try:
        fileString = csTemplate.read()
    finally:
        csTemplate.close()
    
    fileString = fileString.replace('{0}', enumString)
    
    csTarget = open('out/code/ResValue/'+name+".cs", 'w', encoding='utf-8')
    csTarget.write(fileString)
    csTarget.close()

#按行写入CSV文件
def writeCSVFile(file,line,isLast):
    file.writelines(line)
    if isLast :
        file.writelines('\n')

#通过模板生成C#文件
def writeCSharpFile(dec,name,attr,method,hasKey):
    csTemplate = open(TEMPLATE_FILE_NAME, 'r')
    fileString = ""
    try:
        fileString = csTemplate.read()
    finally:
        csTemplate.close()
        
    fileString = fileString.replace('{0}', dec)
    fileString = fileString.replace('{1}', name)
    fileString = fileString.replace('{2}', attr)  
    fileString = fileString.replace('{3}', method)
    fileString = fileString.replace('{4}', hasKey)

    csTarget = open('out/code/ResType/'+name+".cs", 'w', encoding='utf-8')
    csTarget.write(fileString)
    csTarget.close()
    
#  ----------------   添加新的模板C#生成    ---------------------
#通过模板生成C#文件
def writeCSharpLineFile(dec,name):
    csTemplate = open(TEMPLATE_LINE_NAME, 'r')
    fileString = ""
    try:
        fileString = csTemplate.read()
    finally:
        csTemplate.close()
        
    fileString = fileString.replace('{0}', dec)
    fileString = fileString.replace('{1}', name)
    
    csTarget = open('out/code/ResType/'+name+".cs", 'w', encoding='utf-8')
    csTarget.write(fileString)
    csTarget.close()

    #通过模板生成C#文件
def writeCSharpLetterFile(dec,name):
    csTemplate = open(TEMPLATE_LETTER_NAME, 'r')
    fileString = ""
    try:
        fileString = csTemplate.read()
    finally:
        csTemplate.close()
        
    fileString = fileString.replace('{0}', dec)
    fileString = fileString.replace('{1}', name)
    
    csTarget = open('out/code/ResType/'+name+".cs", 'w', encoding='utf-8')
    csTarget.write(fileString)
    csTarget.close()
    
#获取数据表格类型,小于0代表不是数据表
def getDataSheetType(worksheet):
    res = -1
    key1 = worksheet.cell(row=1, column=1).value
    key2 = worksheet.cell(row=1, column=3).value
    key3 = worksheet.cell(row=1, column=5).value
    if key1 == "类型" and key2 == "名称" and key3 == "描述":
        res = worksheet.cell(row=1, column=2).value
    return res

def main(filename):   
    readExcel(filename)
    
if __name__=="__main__":
    if sys.argv[0] != "ExcelLoader.py" :
        os.chdir(sys.argv[0].replace("ExcelLoader.py",""))
    #print(sys.argv[2])
    #print(sys.argv.__len__) 
    if len(sys.argv)>=3:
        cmd = sys.argv[2]
        if cmd == "-data":
            covData = True
            main(sys.argv[1])
        elif cmd == "-code":
            ConvertCommonType.main()
            covCode = True
            main(sys.argv[1])
        elif cmd == "-all":
            ConvertCommonType.main()
            covData = True
            covCode = True
            main(sys.argv[1])
        else:
            print('请输入正确的指令：')
            print('    [filename] -data  生成资源文件')
            print('    [filename] -code  生成代码')
            print('    [filename] -all  全部生成')
    else:
        print('请输入正确的指令：')
        print('    [filename] -data  生成资源文件')
        print('    [filename] -code  生成代码')
        print('    [filename] -all  全部生成')
